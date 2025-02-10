import os
import aiofiles
import csv
import textract
from typing import List
from fastapi import UploadFile, HTTPException
from docx import Document
from openpyxl import load_workbook  # ✅ Import for XLSX processing
from utils import extract_text_from_pdf

UPLOAD_FOLDER = "./uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

async def upload_files(files: List[UploadFile]) -> dict:
    """
    Upload and extract text from multiple file types:
    - PDF, TXT, DOC, DOCX, PAGES, RTF, MD, CSV, and XLSX.
    """
    combined_text = ""
    uploaded_file_types = []

    for file in files:
        try:
            file_extension = file.filename.split(".")[-1].lower()
            file_path = os.path.join(UPLOAD_FOLDER, file.filename)

            # Save file to disk
            async with aiofiles.open(file_path, "wb") as f:
                await f.write(await file.read())

            # Process based on file type
            if file_extension == "pdf":
                pdf_text = extract_text_from_pdf(file_path)
                combined_text += pdf_text + "\n"
                uploaded_file_types.append("PDF")

            elif file_extension in ["txt", "md"]:
                async with aiofiles.open(file_path, "r") as f:
                    txt_text = await f.read()
                combined_text += txt_text + "\n"
                uploaded_file_types.append("Text")

            elif file_extension == "docx":
                doc = Document(file_path)
                docx_text = "\n".join([p.text for p in doc.paragraphs])
                combined_text += docx_text + "\n"
                uploaded_file_types.append("DOCX")

            elif file_extension in ["doc", "pages", "rtf"]:
                textract_text = textract.process(file_path).decode("utf-8")
                combined_text += textract_text + "\n"
                uploaded_file_types.append(file_extension.upper())

            elif file_extension == "csv":
                async with aiofiles.open(file_path, "r") as f:
                    content = await f.read()
                csv_reader = csv.reader(content.splitlines())
                csv_text = "\n".join([", ".join(row) for row in csv_reader])
                combined_text += csv_text + "\n"
                uploaded_file_types.append("CSV")

            elif file_extension == "xlsx":
                wb = load_workbook(file_path, data_only=True)  # ✅ Load Excel workbook
                xlsx_text = ""

                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    xlsx_text += f"\n[Sheet: {sheet}]\n"
                    for row in ws.iter_rows(values_only=True):
                        row_text = ", ".join(str(cell) if cell is not None else "" for cell in row)
                        xlsx_text += row_text + "\n"

                combined_text += xlsx_text + "\n"
                uploaded_file_types.append("XLSX")

            else:
                raise HTTPException(
                    status_code=400, detail=f"Unsupported file type: {file.filename}"
                )

        except Exception as e:
            raise HTTPException(
                status_code=500, detail=f"Error processing {file.filename}: {str(e)}"
            )

    return {
        "message": f"{', '.join(uploaded_file_types)} uploaded successfully",
        "content": combined_text,
    }
